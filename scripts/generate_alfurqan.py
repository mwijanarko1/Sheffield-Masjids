"""
Generate monthly prayer time JSON files for Al Furqan Islamic Centre
from the 2026 timetable Excel file.
"""
import openpyxl
import json
import os
import re
from datetime import datetime, time, timedelta
from collections import defaultdict, Counter

BASE_DIR = "/Users/mikhail/Documents/CURSOR CODES/Deployed/Sheffield-Masjids"
MOSQUE_DIR = os.path.join(BASE_DIR, "public/data/mosques/gb/manchester/alfurqan-islamic-centre")
XLSX_PATH = "/Users/mikhail/Downloads/prayer_times_2026.xlsx"

MONTH_ORDER = [
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"
]


def time_to_str(t):
    """Convert time or datetime.time to HH:MM string."""
    if t is None:
        return None
    if isinstance(t, datetime):
        return t.strftime("%H:%M")
    if isinstance(t, time):
        return t.strftime("%H:%M")
    return str(t)


def add_minutes(t, minutes=5):
    """Add minutes to a time object."""
    if t is None:
        return None
    dt = datetime(2026, 1, 1, t.hour, t.minute, t.second)
    dt += timedelta(minutes=minutes)
    return dt.time()


def parse_maghrib_iqamah(ws, ws_formula, row_idx, maghrib_adhan):
    """
    Determine Maghrib iqamah time.
    - If hardcoded time → use it
    - If formula =G{n}+TIME(0,5,0) → adhan + 5min
    - If formula =G{n}+TIME(0,10,0) → adhan + 10min
    """
    formula_val = ws_formula.cell(row_idx, 12).value
    cached_val = ws.cell(row_idx, 12).value

    if isinstance(formula_val, str) and formula_val.startswith("="):
        # It's a formula — extract the minutes offset
        m = re.search(r'TIME\(0,(\d+),0\)', formula_val)
        if m:
            offset = int(m.group(1))
            return add_minutes(maghrib_adhan, offset)
        # Fallback: +5min
        return add_minutes(maghrib_adhan, 5)
    else:
        # Hardcoded time value — use it directly
        if isinstance(formula_val, time):
            return formula_val
        if isinstance(cached_val, time):
            return cached_val
        # Last resort: +5min
        return add_minutes(maghrib_adhan, 5)


# Load workbook (data_only=True for cached values, data_only=False for formulas)
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active
wb2 = openpyxl.load_workbook(XLSX_PATH, data_only=False)
ws2 = wb2.active

# Column mapping:
# Col 2: GregorianDate (datetime)
# Col 3: FAJR adhan (time)
# Col 4: SUNRISE / Shurooq (time)
# Col 5: DHUHR adhan (time)
# Col 6: ASR adhan (time)
# Col 7: MAGHRIB adhan (time)
# Col 8: ISHA adhan (time)
# Col 9: FAJR iqamah (time)
# Col 10: DHUHR iqamah (time)
# Col 11: ASR iqamah (time)
# Col 12: MAGHRIB iqamah (formula or hardcoded time)
# Col 13: ISHA iqamah (time)
# Col 14: 2ND JUMMAH (time)
# Col 15: 1ST JUMMAH (time)

# Parse all data rows
all_rows = []
for row_idx in range(2, ws.max_row + 1):
    date_val = ws.cell(row_idx, 2).value
    if date_val is None or not isinstance(date_val, datetime):
        continue
    # Only include 2026 data
    if date_val.year != 2026:
        continue

    fajr_adhan = ws.cell(row_idx, 3).value
    shurooq = ws.cell(row_idx, 4).value
    dhuhr_adhan = ws.cell(row_idx, 5).value
    asr_adhan = ws.cell(row_idx, 6).value
    maghrib_adhan = ws.cell(row_idx, 7).value
    isha_adhan = ws.cell(row_idx, 8).value

    fajr_iqamah = ws.cell(row_idx, 9).value
    dhuhr_iqamah = ws.cell(row_idx, 10).value
    asr_iqamah = ws.cell(row_idx, 11).value

    # Maghrib iqamah: handle formulas properly
    maghrib_iqamah = parse_maghrib_iqamah(ws, ws2, row_idx, maghrib_adhan)

    isha_iqamah = ws.cell(row_idx, 13).value

    jummah_2nd = ws.cell(row_idx, 14).value  # 2ND JUMMAH
    jummah_1st = ws.cell(row_idx, 15).value  # 1ST JUMMAH

    all_rows.append({
        "date": date_val,
        "day": date_val.day,
        "month": date_val.month,
        "month_name": date_val.strftime("%B").upper(),
        "fajr": time_to_str(fajr_adhan),
        "shurooq": time_to_str(shurooq),
        "dhuhr": time_to_str(dhuhr_adhan),
        "asr": time_to_str(asr_adhan),
        "maghrib": time_to_str(maghrib_adhan),
        "isha": time_to_str(isha_adhan),
        "fajr_iqamah": time_to_str(fajr_iqamah),
        "dhuhr_iqamah": time_to_str(dhuhr_iqamah),
        "asr_iqamah": time_to_str(asr_iqamah),
        "maghrib_iqamah": time_to_str(maghrib_iqamah),
        "isha_iqamah": time_to_str(isha_iqamah),
        "jummah_1st": time_to_str(jummah_1st),
        "jummah_2nd": time_to_str(jummah_2nd),
        "day_of_week": date_val.strftime("%A"),
    })

print(f"Total 2026 rows: {len(all_rows)}")
print(f"Date range: {all_rows[0]['date']} to {all_rows[-1]['date']}")

# Group by month
months_data = defaultdict(list)
for r in all_rows:
    months_data[r["month_name"]].append(r)

# Generate monthly JSON files
for month_name in MONTH_ORDER:
    if month_name not in months_data:
        print(f"WARNING: No data for {month_name}")
        continue

    entries = months_data[month_name]
    entries.sort(key=lambda x: x["day"])

    print(f"\n{month_name} ({len(entries)} days):")
    print(f"  First: day={entries[0]['day']}, fajr={entries[0]['fajr']}")
    print(f"  Last:  day={entries[-1]['day']}, fajr={entries[-1]['fajr']}")

    # Build prayer_times array
    prayer_times = []
    for e in entries:
        prayer_times.append({
            "date": e["day"],
            "fajr": e["fajr"],
            "shurooq": e["shurooq"],
            "dhuhr": e["dhuhr"],
            "asr": e["asr"],
            "maghrib": e["maghrib"],
            "isha": e["isha"],
        })

    # Build iqamah_times array (one per day)
    iqamah_times = []
    for e in entries:
        iqamah_times.append({
            "date_range": str(e["day"]),
            "fajr": e["fajr_iqamah"],
            "dhuhr": e["dhuhr_iqamah"],
            "asr": e["asr_iqamah"],
            "maghrib": e["maghrib_iqamah"],
            "isha": e["isha_iqamah"],
        })

    # Jummah iqamah: use the most common pair (some months have mid-month transitions)
    from collections import Counter
    jummah_pairs = [(e["jummah_1st"], e["jummah_2nd"]) for e in entries if e["jummah_1st"] and e["jummah_2nd"]]
    if jummah_pairs:
        # Count occurrences of each pair and pick the most common
        pair_counts = Counter(jummah_pairs)
        most_common_pair = pair_counts.most_common(1)[0][0]
        jummah_1st_val, jummah_2nd_val = most_common_pair
        if jummah_2nd_val and jummah_2nd_val != jummah_1st_val:
            jummah_iqamah = f"{jummah_1st_val} / {jummah_2nd_val}"
        else:
            jummah_iqamah = jummah_1st_val
    else:
        jummah_iqamah = ""

    print(f"  Jummah: {jummah_iqamah}")

    # Build JSON
    month_json = {
        "month": month_name,
        "prayer_times": prayer_times,
        "iqamah_times": iqamah_times,
        "jummah_iqamah": jummah_iqamah,
    }

    # Write file
    filename = month_name.lower() + ".json"
    filepath = os.path.join(MOSQUE_DIR, filename)

    with open(filepath, "w") as f:
        json.dump(month_json, f, indent=2)
        f.write("\n")

    print(f"  Written: {filepath}")

print("\nDone! All 12 months generated.")
